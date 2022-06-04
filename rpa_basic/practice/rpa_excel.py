# """ 1 """
# from openpyxl import Workbook

# wb = Workbook() # 새 워크북 생성
# ws = wb.active  # 현재 활성화된 sheet 가져옴
# ws.title = "SheetTitle" # sheet 의 이름을 변경
# wb.save("sample01.xlsx")
# wb.close()

# """ 2 """
# from openpyxl import Workbook

# wb = Workbook() # 새 워크북 생성

# ws = wb.create_sheet()  # 새로운 시트를 기본 이름으로 생성
# ws.title = "MySheet" # 시트 이름 변경
# ws.sheet_properties.tabColor = "ff66ff" # RGB 형태로 시트 탭 색상 변경

# ws1 = wb.create_sheet("YourSheet")  # 새로운 시트를 설정한 이름으로 생성
# # 현재 순서: Sheet, MySheet, YourSheet
# ws2 = wb.create_sheet("NewSheet", 2) # 2번째 인덱스에 시트 생성, My Your 사이

# """
# 작업할 때 
# ws1 = ... 이렇게 작업할 수도 있지만
# wb["NewSheet"] = ... 이렇게 작업할 수도 있다
# """

# print(wb.sheetnames)

# new_ws = wb["NewSheet"]
# new_ws["A1"] = "Test"
# target = wb.copy_worksheet(new_ws)
# target.title = "Copied Sheet"

# print(wb.sheetnames)

# wb.save("sample02.xlsx")
# wb.close()

# """ 3 """
# from openpyxl import Workbook

# wb = Workbook() # 새 워크북 생성
# ws = wb.active
# ws.title = "SheetTitle"

# ws["A1"] = 1 # A1 셀에 1 값을 입력
# ws["A2"] = 2
# ws["A3"] = 3

# ws["B1"] = 4
# ws["B2"] = 5
# ws["B3"] = 6

# print(ws["A1"]) # A1 셀의 정보를 출력
# print(ws["B3"].value)
# print(ws["B3"].row)
# print(ws["B3"].column)
# print(ws["A10"].value)

# ws.cell(row=3, column=3, value=7)
# print(ws["C3"].value)

# ws.cell(4, 4, 8) # row, column, value 순서
# print(ws["D4"].value)

# ws.cell(4, 1, None) # row, column, value 순서
# print(ws["A4"].value)

# from random import *
# index = 1
# for x in range(1, 11): # 10개 row
#     for y in range(1, 11): # 10개 column
#         ws.cell(row=x, column=y, value=(randint(0, 100)))
#         ws.cell(row=x, column=y, value=index)
#         index += 1

# wb.save("sample03.xlsx")

# """ 4 """
# from openpyxl import load_workbook
# wb = load_workbook("sample03.xlsx")
# ws = wb.active

# # cell 데이터 불러오기
# for x in range(1, 11):
#     for y in range(1, 11):
#         print(ws.cell(row=x, column=y).value, end=" ")
#     print()

# # cell 개수를 모를 때
# for x in range(1, ws.max_row+1):
#     for y in range(1, ws.max_column+1):
#         print(ws.cell(row=x, column=y).value, end=" ")
#     print()

# """ 5 """
# from openpyxl import Workbook
# from random import *
# wb = Workbook()
# ws = wb.active # 활성화된 거 가져오기

# # 1줄 씩 데이터 넣기
# ws.append(["번호", "영어", "수학"]) # A, B, C
# for i in range(1, 11): # 10개 데이터
#     ws.append([i, randint(0, 100), randint(0, 100)])

# col_B = ws["B"] # column B 영어
# # print(col_B)
# # for cell in col_B:
# #     print(cell.value)

# col_range = ws["B:C"] # B부터 C까지
# # for cols in col_range:
# #     for cell in cols:
# #         print(cell.value)

# row_title = ws[1] # 1번째 row
# # for cell in row_title:
# #     print(cell.value)

# row_range = ws[2:6] # 2번째 줄부터 6번째 줄까지 가지고 오기 2, 3, 4, 5, 6(6도 포함)
# # for rows in row_range:
# #     for cell in rows:
# #         print(cell.value, end=" ")
# #     print()

# from openpyxl.utils.cell import coordinate_from_string

# row_range = ws[2:ws.max_row] # 2번째 줄부터 마지막 줄까지
# # for rows in row_range:
# #     for cell in rows:
# #         # print(cell.value, end=" ")
# #         # print(cell.coordinate, end=" ") # cell 번호 출력
# #         xy = coordinate_from_string(cell.coordinate)
# #         # print(xy, end=" ")
# #         print(xy[0], xy[1], end=" ")
# #     print()

# # # 전체 rows
# # print(tuple(ws.rows))
# # for row in tuple(ws.rows):
# #     print(row[2].value)

# # # 전체 columns
# # print(tuple(ws.columns))
# # for column in tuple(ws.columns):
# #     print(column[0].value)

# # for row in ws.iter_rows(): # 전체 row
# #     print(row[2].value)

# # for column in ws.iter_cols(): # 전체 column
# #     print(column[0].value)

# # # 2번째 줄부터 11번째 줄까지, 2번째 열부터 3번째 열까지
# # for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):
# #     # print(row[0].value, row[1].value) # 수학, 영어
# #     print(row)

# # for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):
# #     print(col)

# wb.save("sample05.xlsx")

# """ 6 """
# from openpyxl import load_workbook
# wb = load_workbook("sample05.xlsx")
# ws = wb.active # 활성화된 거 가져오기

# for row in ws.iter_rows(min_row=2):
#     # 번호, 영어, 수학
#     if int(row[1].value) > 80:
#         print(row[0].value, "번 학생은 영어 천재")

# for row in ws.iter_rows(max_row=1):
#     for cell in row:
#         if cell.value == "영어":
#             cell.value = "컴퓨터"

# wb.save("sample06.xlsx")

# """ 7 """
# from openpyxl import load_workbook
# wb = load_workbook("sample05.xlsx")
# ws = wb.active # 활성화된 거 가져오기

# # ws.insert_rows(8) # 8번째 줄에 삽입
# # ws.insert_rows(8, 5) # 8번째 줄부터 다섯 줄 삽입
# # ws.insert_cols(2)
# ws.insert_cols(2, 3)

# wb.save("sample07.xlsx")

# """ 8 """
# from openpyxl import load_workbook
# wb = load_workbook("sample05.xlsx")
# ws = wb.active # 활성화된 거 가져오기

# # ws.delete_rows(8) # 8번째 줄에 있는 7번 학생 데이터 삭제
# # ws.delete_rows(8, 3) # 8번째 줄 부터 세 줄 삭제
# # ws.delete_cols(2) # 두 번째 열 삭제(B 삭제)
# ws.delete_cols(2, 2)

# wb.save("sample08.xlsx")

# """ 9 """
# from openpyxl import load_workbook
# wb = load_workbook("sample05.xlsx")
# ws = wb.active # 활성화된 거 가져오기

# # # 번호 영어 수학 -> 번호 국어 영어 수학
# # ws.move_range("B1:C11", rows=0, cols=1) # ""범위를 row방향 0칸, col방향 1칸 이동
# # ws["B1"].value = "국어"

# ws.move_range("C1:C11", rows=5, col=-1)

# wb.save("sample09.xlsx")

# """ 10 """
# from openpyxl import load_workbook
# wb = load_workbook("sample05.xlsx")
# ws = wb.active # 활성화된 거 가져오기

# from openpyxl.chart import BarChart, Reference, LineChart
# # B2:C11까지의 데이터를 차트로 생성
# # bar_value = Reference(ws, min_row=2, max_row=11, min_col=2, max_col=3)
# # bar_chart = BarChart()
# # bar_chart.add_data(bar_value)

# # ws.add_chart(bar_chart, "E1") # 차트 넣을 위치 정의

# # B1:C11까지의 데이터를 차트로 생성
# line_value = Reference(ws, min_row=1, max_row=11, min_col=2, max_col=3)
# line_chart = LineChart()
# line_chart.add_data(line_value, titles_from_data=True) # 계열 > 영어, 수학
# line_chart.title = "성적표" # 제목
# line_chart.style = 10 # 미리 정의된 스타일 적용, 사용자가 개별 지정 가능
# line_chart.y_axis.title = "점수" # Y축 제목
# line_chart.x_axis.title = "번호"
# ws.add_chart(line_chart, "E1")

# wb.save("sample10.xlsx")

# """ 11 """
# from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
# from openpyxl import load_workbook
# wb = load_workbook("sample05.xlsx")
# ws = wb.active # 활성화된 거 가져오기

# # 번호, 영어, 수학
# a1 = ws["A1"]
# b1 = ws["B1"]
# c1 = ws["C1"]

# # A열의 너비를 5로 설정
# ws.column_dimensions["A"].width = 5

# # 1행의 높이를 50으로 설정
# ws.row_dimensions[1].height = 50

# # 스타일 적용
# a1.font = Font(color="FF0000", italic=True, bold=True) # "RGB", 이탤릭, 볼드체
# b1.font = Font(color="CC33FF", name="Arial", strike=True)
# c1.font = Font(color="0000FF", size=20, underline="single")

# # 테두리 적용
# thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
# a1.border = thin_border
# b1.border = thin_border
# c1.border = thin_border

# # 90점 이상인 셀에 대해 초록색 배경 적용
# for row in ws.rows:
#     for cell in row:
#         # 각 cell에 대해 정렬, center, left, top, bottom, right
#         cell.alignment = Alignment(horizontal="center", vertical="center")
#         if cell.column == 1: # A열은 제외
#             continue
#         # cell 이 정수형 데이터이고 90점 이상이면
#         if isinstance(cell.value, int) and cell.value >= 90:
#             cell.fill = PatternFill(fgColor="00FF00", fill_type="solid")
#             cell.font = Font(color="FF0000")

# # 틀 고정
# ws.freeze_panes = "B2" # B2 기준 틀 고정

# wb.save("sample11.xlsx")

# """ 12 """
# import datetime
# from openpyxl import Workbook
# wb = Workbook()
# ws = wb.active

# ws["A1"] = datetime.datetime.today() # 오늘 날짜 정보
# ws["A2"] = "=SUM(1, 2, 3)" # 1+2+3 = 6
# ws["A3"] = "=AVERAGE(1, 2, 3)" # 평균 2

# ws["A4"] = 10
# ws["A5"] = 20
# ws["A6"] = "=SUM(A4:A5)"

# wb.save("sample12.xlsx")

""" 13 """
from openpyxl import load_workbook
wb = load_workbook("sample12.xlsx", data_only=True)
ws = wb.active # 활성화된 거 가져오기

for row in ws.values: # .values: 값을 바로 가져옴
    for cell in row:
        print(cell) # dataonly가 없을 경우 수식 그대로 가져옴 2022-05-18 19:30:56.593000, =SUM(1, 2, 3), =AVERAGE(1, 2, 3)...



# wb.save("sample13.xlsx")